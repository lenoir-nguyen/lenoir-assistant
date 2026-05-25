"""
Document processor for parsing and chunking various file formats.

Supports: PDF, TXT, Markdown, Word (.docx), Excel (.xlsx), PNG, JPEG
Extensible design for adding more formats in the future.
"""

import re
from pathlib import Path
from typing import List
from config import get_settings

settings = get_settings()


class DocumentProcessor:
    """Process documents and split into chunks for embedding."""

    CHUNK_SIZE = settings.CHUNK_SIZE or 1024  # tokens per chunk
    CHUNK_OVERLAP = settings.CHUNK_OVERLAP or 100  # tokens overlap

    @staticmethod
    async def process_document(file_path: str, file_extension: str) -> str:
        """
        Process document based on file type and extract text.

        Args:
            file_path: Path to the file on disk
            file_extension: File extension (.pdf, .txt, .docx, etc.)

        Returns:
            Extracted text content

        Raises:
            ValueError: If file format not supported
        """
        ext = file_extension.lower()

        if ext == '.pdf':
            return await DocumentProcessor.process_pdf(file_path)
        elif ext == '.txt':
            return await DocumentProcessor.process_txt(file_path)
        elif ext == '.md':
            return await DocumentProcessor.process_markdown(file_path)
        elif ext == '.docx':
            return await DocumentProcessor.process_word(file_path)
        elif ext == '.xlsx':
            return await DocumentProcessor.process_excel(file_path)
        elif ext in ['.png', '.jpg', '.jpeg']:
            return await DocumentProcessor.process_image(file_path)
        else:
            raise ValueError(f"Unsupported file format: {ext}")

    @staticmethod
    async def process_pdf(file_path: str) -> str:
        """Extract text from PDF file."""
        try:
            import pdfplumber
        except ImportError:
            raise ImportError("pdfplumber required for PDF processing. Install with: pip install pdfplumber")

        import asyncio
        loop = asyncio.get_event_loop()

        text = await loop.run_in_executor(
            None,
            lambda: DocumentProcessor._extract_pdf_text(file_path)
        )
        return text

    @staticmethod
    def _extract_pdf_text(file_path: str) -> str:
        """Helper to extract text from PDF (blocking operation)."""
        import pdfplumber
        text = []
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text.append(page_text)
        return "\n".join(text)

    @staticmethod
    async def process_txt(file_path: str) -> str:
        """Extract text from plain text file."""
        import asyncio
        loop = asyncio.get_event_loop()

        return await loop.run_in_executor(
            None,
            lambda: open(file_path, 'r', encoding='utf-8').read()
        )

    @staticmethod
    async def process_markdown(file_path: str) -> str:
        """Extract text from Markdown file."""
        import asyncio
        loop = asyncio.get_event_loop()

        return await loop.run_in_executor(
            None,
            lambda: open(file_path, 'r', encoding='utf-8').read()
        )

    @staticmethod
    async def process_word(file_path: str) -> str:
        """Extract text from Word (.docx) file."""
        try:
            from docx import Document
        except ImportError:
            raise ImportError("python-docx required for Word processing. Install with: pip install python-docx")

        import asyncio
        loop = asyncio.get_event_loop()

        text = await loop.run_in_executor(
            None,
            lambda: DocumentProcessor._extract_docx_text(file_path)
        )
        return text

    @staticmethod
    def _extract_docx_text(file_path: str) -> str:
        """Helper to extract text from Word doc (blocking operation)."""
        from docx import Document
        doc = Document(file_path)
        return "\n".join(paragraph.text for paragraph in doc.paragraphs)

    @staticmethod
    async def process_excel(file_path: str) -> str:
        """Extract text from Excel (.xlsx) file."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl required for Excel processing. Install with: pip install openpyxl")

        import asyncio
        loop = asyncio.get_event_loop()

        text = await loop.run_in_executor(
            None,
            lambda: DocumentProcessor._extract_xlsx_text(file_path)
        )
        return text

    @staticmethod
    def _extract_xlsx_text(file_path: str) -> str:
        """Helper to extract text from Excel file (blocking operation)."""
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
        text = []

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            text.append(f"Sheet: {sheet}")
            for row in ws.iter_rows(values_only=True):
                # Filter out None values and join
                row_text = " | ".join(str(cell) for cell in row if cell is not None)
                if row_text.strip():
                    text.append(row_text)

        return "\n".join(text)

    @staticmethod
    async def process_image(file_path: str) -> str:
        """Extract text from image (PNG/JPEG) using OCR."""
        try:
            import pytesseract
            from PIL import Image
        except ImportError:
            raise ImportError("pytesseract and Pillow required for image OCR. Install with: pip install pytesseract pillow")

        import asyncio
        loop = asyncio.get_event_loop()

        text = await loop.run_in_executor(
            None,
            lambda: DocumentProcessor._extract_image_text(file_path)
        )
        return text

    @staticmethod
    def _extract_image_text(file_path: str) -> str:
        """Helper to extract text from image (blocking operation)."""
        import pytesseract
        from PIL import Image

        img = Image.open(file_path)
        return pytesseract.image_to_string(img)

    @staticmethod
    def chunk_text(text: str, chunk_size: int = None, chunk_overlap: int = None) -> List[str]:
        """
        Split text into overlapping chunks.

        Args:
            text: Full document text
            chunk_size: Approximate tokens per chunk (default: CHUNK_SIZE)
            chunk_overlap: Tokens to overlap between chunks (default: CHUNK_OVERLAP)

        Returns:
            List of text chunks
        """
        chunk_size = chunk_size or DocumentProcessor.CHUNK_SIZE
        chunk_overlap = chunk_overlap or DocumentProcessor.CHUNK_OVERLAP

        # Simple chunking by character count (rough estimate of tokens)
        # 1 token ≈ 4 characters as rough approximation
        char_size = chunk_size * 4
        char_overlap = chunk_overlap * 4

        chunks = []
        start = 0

        while start < len(text):
            end = min(start + char_size, len(text))
            chunk = text[start:end]

            # Try to break at sentence boundary if possible
            if end < len(text):
                # Look for last period/newline within chunk
                last_sentence = max(
                    chunk.rfind('. '),
                    chunk.rfind('\n'),
                    chunk.rfind('! '),
                    chunk.rfind('? ')
                )
                if last_sentence > char_size * 0.5:  # At least 50% through chunk
                    end = start + last_sentence + 1

            chunk = text[start:end].strip()
            if chunk:  # Only add non-empty chunks
                chunks.append(chunk)

            # Move forward, leaving overlap
            start = end - char_overlap

        return chunks

    @staticmethod
    def get_supported_formats() -> List[str]:
        """Return list of supported file extensions."""
        return ['.pdf', '.txt', '.md', '.docx', '.xlsx', '.png', '.jpg', '.jpeg']
